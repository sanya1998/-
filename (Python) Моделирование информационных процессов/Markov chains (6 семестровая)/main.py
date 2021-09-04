import numpy as np
from openpyxl import Workbook

# Порядковое число
number_in_list = 9
# Число этапов
N = 3
# Число состояний
m = 3
# Коэффициент дисконтирования
gamma = number_in_list / (number_in_list + 1)
print("gamma =", gamma)
# gamma = 1

# Список матриц вероятностей перехода из i в j состояние
P = [
		[
			[0.2, 0.5, 0.3],
			[0, 0.5, 0.5],
			[0, 0, 1]
		],
		[
			[0.3, 0.6, 0.1],
			[0.1, 0.6, 0.3],
			[0.05, 0.4, 0.55]
		]
		# [
		# 	[0.2, 0.5, 0.3],
 		#  [0, 0.5, 0.5],
		# 	[0.05, 0.4, 0.55]
		# ]
		# [
		# 	[0.2, 0.5, 0.3],
 		# 	[0.1, 0.6, 0.3],
		# 	[0.05, 0.4, 0.55]
		# ]
	]
# Число альтернатив
K = len(P)
# Список матриц доходов перехода из i в j состояние
A = [
		[
			[7, 6, 3],
			[0, 5, 1],
			[0, 0, -1]
		],
		[
			[6, 5, -1],
			[7, 4, 0],
			[6, 3, -2]
		]
		# [
		# 	[7, 6, 3],
		#  [0, 5, 1],
		# 	[6, 3, -2]
		# ]
		# [
		# 	[7, 6, 3],
		#  	[7, 4, 0],
		# 	[6, 3, -2]
		# ]
	]
# Запись в excel
wb = Workbook()
ws = wb.active
for k in range(K):
	for i in range(m):
		ws.append(P[k][i] + [' '] + A[k][i])
	ws.append([])

P = np.array(P)
A = np.array(A)

row_in_ws = (K+1)*m
# Подсчет b
b = np.zeros(shape=(K, m))
ws.cell(column=1, row=row_in_ws).value = 'i'
for i in range(m):
	ws.cell(column=1, row=row_in_ws+i+1).value = i+1
	for k in range(K):
		# Выполнится несколько раз, ну ладно...
		ws.cell(column=k+2, row=row_in_ws).value = 'b_'+str(k+1)+'_i'
		b[k][i] = P[k][i] @ A[k][i]
		ws.cell(column=k+2, row=row_in_ws + i + 1).value = b[k][i]

# Последовательности альтернатив
sequence_alternatives = list()
# Подсчет значений функций
f_prev = np.zeros(m)
for n in range(N):
	f_next = np.zeros(m)
	k_star = np.zeros(m, dtype=int)
	row_in_ws += m + 2
	ws.cell(column=1, row=row_in_ws).value = 'i'
	ws.cell(column=K + 2, row=row_in_ws).value = 'f' + str(n + 1) + '(i)'
	ws.cell(column=K + 3, row=row_in_ws).value = 'k*'
	for i in range(m):
		ws.cell(column=1, row=row_in_ws + i + 1).value = i + 1
		val_max = -10**10
		k_argmax = 0
		for k in range(K):
			# Выполнится несколько раз, ну ладно...
			ws.cell(column=k + 2, row=row_in_ws).value = 'b_' + str(k+1) + '_i + gamma*sum_po_j( p_i_j*f_' + str(n+1) + '(j) )'
			val = b[k][i] + gamma*(P[k][i] @ f_prev)
			ws.cell(column=k + 2, row=row_in_ws + i + 1).value = val
			if val > val_max:
				val_max = val
				k_argmax = k
		f_next[i] = val_max
		ws.cell(column=K + 2, row=row_in_ws+i+1).value = f_next[i]
		k_star[i] = k_argmax + 1
		ws.cell(column=K + 3, row=row_in_ws+i+1).value = k_star[i]
	sequence_alternatives.append(k_star)
	print(f_next, k_star)
	f_prev = f_next

# Внесение последовательности альтернатив таблицу
row_in_ws += m + 2
ws.cell(column=1, row=row_in_ws).value = 'Этапы'
for n in range(N):
	ws.cell(column=n+2, row=row_in_ws).value = n+1
	for i in range(m):
		ws.cell(column=n+2, row=row_in_ws+i+1).value = 'i=' + str(i+1) + ' -> k=' + \
		                                               str(sequence_alternatives[N-n-1][i])
wb.save('sample.xlsx')
# print(sequence_alternatives)
print("Все необходмые данные в excel")
