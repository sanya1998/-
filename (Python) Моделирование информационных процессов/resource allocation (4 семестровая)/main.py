from openpyxl import Workbook
import numpy as np

# Генерация f_table
# f_table = [[0,8,10,11,12,18],
#            [0,6,9,11,13,15],
#            [0,3,4,7,11,18],
#            [0,4,6,8,13,16]]
f_table = list()
max_in_col = 0
for nom in range(6):
	line = list()
	if nom == 0:
		for nom2 in range(4):
			line.append(0)
	elif nom == 1:
		max_in_col = k = 9  # Моё порядковое число
		for nom2 in range(3):
			number = np.random.randint(k-2, k+3)
			line.append(number)
			if number > max_in_col:
				max_in_col = number
		line.append(k)
	else:
		k = max_in_col
		for nom2 in range(4):
			number = np.random.randint(k+1, k + 6)
			line.append(number)
			if number > max_in_col:
				max_in_col = number
		
	f_table.append(line)
	
f_table = np.array(f_table).T
print(f_table)

# Запись в excel
wb = Workbook()
ws = wb.active
# Заносим таблицу маленьких f
ws.append(["x"] + list(np.arange(6)*2))
for i, line in enumerate(f_table):
	ws.append(["f"+str(i+1) + "(x)"] + list(line))

row = 7
col = 1
F_prev = list(f_table[len(f_table)-1])
arg_F = list(np.arange(6)*2)
F_table = list()
x_table = list()
for i,f in enumerate(f_table[::-1]):
	# Первая таблица не нужна
	if i != 0:
		# Большие таблицы
		ws.cell(column=col, row=row).value = "s"
		ws.cell(column=col, row=row+1).value = "x"+str(len(f_table)-i)
		# ws.cell(column=col, row=row+2).value = "f" + str(len(f_table)-i) + "(x"+str(len(f_table)-i) + ")+F" + str(len(f_table)-i+1) + "(s-" + "x"+str(len(f_table)-i) + ")"
		ws.cell(column=col, row=row+2).value = "f" + str(len(f_table)-i) + "+F" + str(len(f_table)-i+1)
		col += 1
		F_next = list()
		arg_F = list()
		for s in list(np.arange(6)*2):
			arg_val = 0
			max_val = 0
			for x_i in list(np.arange(s/2+1)*2):
				ws.cell(column=col, row=row).value = s
				ws.cell(column=col, row=row+1).value = x_i
				val = f[int(x_i/2)]+F_prev[int((s-x_i)/2)]
				ws.cell(column=col, row=row+2).value = val
				if val > max_val:
					max_val = val
					arg_val = x_i
				col += 1
			F_next.append(max_val)
			arg_F.append(int(arg_val))
		F_prev = list(F_next)
		row += 4
	F_table.append(F_prev)
	x_table.append(arg_F)
	# Таблица максимумов
	col=1
	ws.cell(column=col, row=row).value = "s"
	ws.cell(column=col, row=row + 1).value = "x" + str(len(f_table) - i)
	# ws.cell(column=col, row=row + 2).value = "F" + str(len(f_table) - i) + "(s)"
	ws.cell(column=col, row=row + 2).value = "F" + str(len(f_table) - i)
	col += 1
	for i,col_F in enumerate(F_prev):
		ws.cell(column=col, row=row).value = i*2
		ws.cell(column=col, row=row + 1).value = arg_F[i]
		ws.cell(column=col, row=row + 2).value = col_F
		col += 1
	row += 4
	col=1

wb.save('sample.xlsx')

# print(F_table)
# print(x_table)

profit = np.max(F_table[len(F_table)-1])
print("Максимум:", profit)
pos = np.argmax(F_table[len(F_table)-1])
s=10
answers_x = list()
for nom, iks in enumerate(x_table[::-1]):
	x = iks[pos]
	print("s=" + str(s) + "=>x" + str(nom+1) + "=" + str(x) + "\ns=s-" + str(x))
	s -= x
	answers_x.append(x)
	pos = int(s/2)
print("x =", answers_x)

check = 0
print("Проверка:")
text_str = ""
for nom, line in enumerate(f_table):
	if nom < len(f_table) - 1:
		text_str += "f" + str(nom+1) + "(" + str(answers_x[nom]) + ") + "
	else:
		text_str += "f" + str(nom + 1) + "(" + str(answers_x[nom]) + ") = "
for nom, line in enumerate(f_table[::-1]):
	if nom < len(f_table) - 1:
		text_str += str(f_table[nom][int(answers_x[nom]/2)]) + " + "
	else:
		text_str += str(f_table[nom][int(answers_x[nom]/2)]) + " = "
	check += f_table[nom][int(answers_x[nom]/2)]
text_str += str(check)
print( text_str)